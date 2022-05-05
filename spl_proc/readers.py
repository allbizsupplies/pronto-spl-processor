
import re
from typing import Any, Dict, List
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from spl_proc.dataclasses import SupplierPricelistItem


DatagridRow = Dict[str, Any]


class DataGridReader:
    HEADER_ROW_INDEX = 1

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.workbook: Workbook = load_workbook(
            filename=self.filepath, read_only=True)

    def load_spl_items_from_worksheet(self, worksheet_name: str | None = None):
        if worksheet_name is None:
            worksheet_name = self.get_default_worksheet_name()
        worksheet: Worksheet = self.workbook.get_sheet_by_name(worksheet_name)
        return self.read_rows_from_worksheet(worksheet)

    def get_default_worksheet_name(self):
        return str(self.workbook.get_sheet_names()[0])

    def read_rows_from_worksheet(self, worksheet: Worksheet):
        rows: List[SupplierPricelistItem] = []
        fieldnames = self.get_fieldnames_from_header_row(worksheet)
        for row_index in self.get_occupied_row_indices(worksheet):
            values = self.get_row_values(worksheet, fieldnames, row_index)
            rows.append(self.convert_row_to_spl_item(values))
        return rows

    def get_row_values(self, worksheet: Worksheet, fieldnames: List[str], row_index: int):
        values: DatagridRow = {}
        for col_index, fieldname in enumerate(fieldnames, start=1):
            value = worksheet.cell(row_index, col_index).value
            values[fieldname] = value
        return values

    def convert_row_to_spl_item(self, row: DatagridRow):
        return SupplierPricelistItem(
            supp_code=row["supp_code"],
            supp_item_code=row["supp_item_code"],
            supp_price=row["supp_price"],
            supp_uom=row["supp_uom"],
            supp_sell_uom=row["supp_sell_uom"],
            supp_eoq=row["supp_eoq"],
            supp_conv_factor=row["supp_conv_factor"])

    def get_fieldnames_from_header_row(self, worksheet: Worksheet):
        fieldnames: List[str] = []
        col_index = 0
        reached_last_column = False
        while not reached_last_column:
            value = worksheet.cell(self.HEADER_ROW_INDEX, col_index + 1).value
            if value is None:
                reached_last_column = True
            else:
                fieldnames.append(self.fieldname_to_snakecase(value))
                col_index += 1
        return fieldnames

    def get_occupied_row_indices(self, worksheet: Worksheet):
        row_indices: List[int] = []
        row_index = self.HEADER_ROW_INDEX + 1
        reached_last_row = False
        while not reached_last_row:
            if worksheet.cell(row_index, 1).value is not None:
                row_indices.append(row_index)
                row_index += 1
            else:
                reached_last_row = True
        return row_indices

    def fieldname_to_snakecase(self, value: str):
        value = value.strip()
        value = value.lower()
        value = re.sub(r"[/ =:-]", "_", value)
        value = re.sub(r"[\(\)\.]", "", value)
        value = value.replace("\n", "_")
        value = value.replace("___", "_")
        value = value.replace("__", "_")
        value = value.strip("_")
        return value


class BRO_craft(DataGridReader):
    HEADER_ROW_INDEX = 2

    def get_occupied_row_indices(self, worksheet: Worksheet):
        row_indices: List[int] = []
        row_index = self.HEADER_ROW_INDEX + 1
        reached_last_row = False
        while not reached_last_row:
            if worksheet.cell(row_index, 1).value is None:
                reached_last_row = True
            elif worksheet.cell(row_index, 2).value is not None:
                row_indices.append(row_index)
            row_index += 1
        return row_indices

    def convert_row_to_spl_item(self, row: DatagridRow):
        return SupplierPricelistItem(
            supp_code="BRO",
            supp_item_code=self.parse_supp_item_code(row["product"]),
            supp_price=row["ex_gst"],
            supp_uom="EACH",
            supp_sell_uom="EACH",
            supp_eoq=1,
            supp_conv_factor=1)

    def parse_supp_item_code(self, value):
        value = value.split("\n")[0]
        value = value.split(":")[0]
        value = value.replace(" NEW", "")
        value = value.strip()
        return value


class CSS(DataGridReader):
    HEADER_ROW_INDEX = 1

    def get_occupied_row_indices(self, worksheet: Worksheet):
        row_indices: List[int] = []
        row_index = self.HEADER_ROW_INDEX + 1
        reached_last_row = False
        while not reached_last_row:
            item_code = worksheet.cell(row_index, 1).value
            if item_code is None:
                reached_last_row = True
            else:
                row_indices.append(row_index)
            row_index += 1
        return row_indices

    def convert_row_to_spl_item(self, row: DatagridRow):
        return SupplierPricelistItem(
            supp_code="CSS",
            supp_item_code=row["code"],
            supp_price=row["buy_price_ex_gst"],
            supp_uom="EACH",
            supp_sell_uom="EACH",
            supp_eoq=1,
            supp_conv_factor=1)

    def get_fieldnames_from_header_row(self, worksheet: Worksheet):
        fieldnames = super().get_fieldnames_from_header_row(worksheet)
        for index, fieldname in enumerate(fieldnames):
            if fieldname.endswith("cssc_sell"):
                fieldnames[index] = "buy_price_ex_gst"
        return fieldnames


class DYN(DataGridReader):
    HEADER_ROW_INDEX = 3

    def get_occupied_row_indices(self, worksheet: Worksheet):
        row_indices: List[int] = []
        row_index = self.HEADER_ROW_INDEX + 1
        reached_last_row = False
        while not reached_last_row:
            description = worksheet.cell(row_index, 4).value
            item_code = worksheet.cell(row_index, 2).value
            buy_price = worksheet.cell(row_index, 6).value
            if description is None:
                reached_last_row = True
            elif (
                item_code is not None and
                item_code != "TBA" and
                buy_price != "POA"
            ):
                row_indices.append(row_index)
            row_index += 1
        return row_indices

    def convert_row_to_spl_item(self, row: DatagridRow):
        return SupplierPricelistItem(
            supp_code="DYN",
            supp_item_code=row["ds_code"],
            supp_price=row["buy_price_ex_gst"],
            supp_uom="EACH",
            supp_sell_uom="EACH",
            supp_eoq=1,
            supp_conv_factor=1)
